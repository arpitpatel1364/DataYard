"""
REPORTING ENGINE

Generates downloadable PDF, CSV, JSON, and Excel reports from a dataset
version's full scan report, including scores, findings and recommendations.
"""
import json
import csv
from pathlib import Path

from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle)
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
import openpyxl
from openpyxl.styles import Font, PatternFill

BRAND_COLOR = "#72ab52"


def _scores_table_data(version: dict) -> list[list]:
    return [
        ["Metric", "Score"],
        ["Health Score", version.get("health_score")],
        ["Health Grade", version.get("health_grade")],
        ["Integrity", version.get("integrity_score")],
        ["Annotation Quality", version.get("annotation_score")],
        ["Balance", version.get("balance_score")],
        ["Image Quality", version.get("image_quality_score")],
        ["Diversity", version.get("diversity_score")],
        ["Leakage", version.get("leakage_score")],
    ]


def generate_pdf_report(dataset_name: str, version: dict, recommendations: list[dict],
                         output_path: Path) -> Path:
    doc = SimpleDocTemplate(str(output_path), pagesize=A4,
                             topMargin=2 * cm, bottomMargin=2 * cm)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("CISTitle", parent=styles["Title"], textColor=colors.HexColor(BRAND_COLOR))
    heading_style = ParagraphStyle("CISHeading", parent=styles["Heading2"], textColor=colors.HexColor(BRAND_COLOR))

    elements = [
        Paragraph("Cactus Intelligence Suite", title_style),
        Paragraph(f"Dataset Health Report - {dataset_name}", styles["Heading3"]),
        Spacer(1, 0.5 * cm),
        Paragraph(f"Generated for version #{version.get('version_number')} "
                  f"(scan mode: {version.get('scan_mode')})", styles["Normal"]),
        Spacer(1, 0.6 * cm),
        Paragraph("Health Scores", heading_style),
    ]

    table = Table(_scores_table_data(version), colWidths=[8 * cm, 6 * cm])
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(BRAND_COLOR)),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.whitesmoke, colors.white]),
    ]))
    elements += [table, Spacer(1, 0.8 * cm), Paragraph("Findings & Recommendations", heading_style)]

    for rec in recommendations:
        sev = rec.get("severity", "info").upper()
        elements.append(Paragraph(f"<b>[{sev}]</b> {rec.get('message', '')}", styles["Normal"]))
        elements.append(Spacer(1, 0.15 * cm))

    doc.build(elements)
    return output_path


def generate_csv_report(version: dict, output_path: Path) -> Path:
    with open(output_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerows(_scores_table_data(version))
    return output_path


def generate_json_report(dataset_name: str, version: dict, recommendations: list[dict],
                          output_path: Path) -> Path:
    payload = {
        "dataset_name": dataset_name,
        "version": version,
        "recommendations": recommendations,
    }
    output_path.write_text(json.dumps(payload, indent=2, default=str))
    return output_path


def generate_excel_report(dataset_name: str, version: dict, recommendations: list[dict],
                           output_path: Path) -> Path:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Health Scores"
    header_fill = PatternFill(start_color="72AB52", end_color="72AB52", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for row in _scores_table_data(version):
        ws.append(row)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 16

    ws2 = wb.create_sheet("Recommendations")
    ws2.append(["Category", "Severity", "Message"])
    for cell in ws2[1]:
        cell.fill = header_fill
        cell.font = header_font
    for rec in recommendations:
        ws2.append([rec.get("category"), rec.get("severity"), rec.get("message")])
    ws2.column_dimensions["A"].width = 22
    ws2.column_dimensions["B"].width = 12
    ws2.column_dimensions["C"].width = 80

    wb.save(output_path)
    return output_path
